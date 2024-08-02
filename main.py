import csv
import json

from dotenv import load_dotenv
import openpyxl
from tabulate import tabulate 
from db.dbconn import SqlServerDB
from db.query import *
from openpyxl import Workbook
import xmltodict
from deepdiff import DeepDiff
import os
import datetime

load_dotenv()
db24 = SqlServerDB(os.getenv('SQL_SERVER_DB_24_SERVER'),os.getenv('SQL_SERVER_DB_24_DATABASE'),os.getenv('SQL_SERVER_DB_24_USERNAME'),os.getenv('SQL_SERVER_DB_24_PASSWORD'))
db44 = SqlServerDB(os.getenv('SQL_SERVER_DB_44_SERVER'),os.getenv('SQL_SERVER_DB_44_DATABASE'),os.getenv('SQL_SERVER_DB_44_USERNAME'),os.getenv('SQL_SERVER_DB_44_PASSWORD'))
querys = Query()


def getAllSchemaByName():
    try:
        print(querys.simData())
        allName = []
        db44.connect()
        data = db44.execute_query(querys.simData())
        for i in range(len(data)-1):
            date_value = data[i].get('ValidFrom')
            allName.append((data[i].get('name'), date_value.strftime('%Y-%m-%d %H:%M:%S')))
    except Exception as e:
        return
    finally:
        db44.close()
    return set(allName)

def loadSchemaInfo(name, datefrom):
    try:
        db24.connect()
        db44.connect()
        data_v3 = db24.execute_query(querys.getgetSchemaByName('bagsPAMF_CBS', name, datefrom))
        data_v4 = db44.execute_query(querys.getgetSchemaByName('bagsPAMF4', name, datefrom))
        return (name, data_v3,data_v4)
    except Exception as e:
        return
    finally:
        db24.close()
        db44.close()

def parse_xml(xml_string):
    return xmltodict.parse(xml_string)

def compare_xml_structures(old_xml, new_xml):
    diff = DeepDiff(old_xml, new_xml, ignore_order=True)
    return diff


def write_to_excel(changes, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = data[i][0]

    # Écrire les en-têtes
    ws['A1'] = "Élément"
    ws['B1'] = "Ancienne valeur"
    ws['C1'] = "Nouvelle valeur"

    # Écrire les changements
    for row, change in enumerate(changes, start=2):
        ws.cell(row=row, column=1, value=change[0])
        ws.cell(row=row, column=2, value=str(change[1]))
        ws.cell(row=row, column=3, value=str(change[2]))

    wb.save(filename)

def copy_structure_only(xml_content):
    data_dict = xmltodict.parse(xml_content)

    def copy_structure(node):
        if isinstance(node, dict):
            return {k: copy_structure(v) for k, v in node.items() if k not in ['#text', '@']}
        elif isinstance(node, list):
            return [copy_structure(item) for item in node]
        else:
            return None

    structure_dict = copy_structure(data_dict)

    def clean_structure(d):
        if isinstance(d, dict):
            return {k: clean_structure(v) for k, v in d.items() if v is not None}
        elif isinstance(d, list):
            return [clean_structure(item) for item in d if item is not None]
        return d
    
    clean_structure_dict = clean_structure(structure_dict)

    structure_xml = xmltodict.unparse(clean_structure_dict, pretty=True)
    
    return structure_xml

def format_path(path):
    """
    Format the path for better readability
    """
    path = path.replace("root", "").strip("['']").replace("']['", " -> ").replace("]['", '->').replace("'][", '->')
    return path

def format_value(value):
    """
    Format the value for better readability
    """
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, indent=2)
    if value is None:
        return ''
    return str(value)

nameListes = getAllSchemaByName()
print(nameListes)
data = []

for element in nameListes:
    data.append(loadSchemaInfo(element[0], element[1]))

for element in nameListes:
    data.append(loadSchemaInfo(element[0], element[1]))

for i in range(len(data)):
    try:
        xml_v3 = data[i][1][0]['CombinationXML']   
    except Exception as e:
        xml_v3 = copy_structure_only(data[i][2][0]['CombinationXML'])
    try:
        xml_v4 = data[i][2][0]['CombinationXML'] 
    except Exception as e:
        xml_v4 = copy_structure_only(data[i][1][0]['CombinationXML'])
    
    old_dict = xmltodict.parse(xml_v3)
    new_dict = xmltodict.parse(xml_v4)
    diff = DeepDiff(old_dict, new_dict)
    diff_data = []
    if 'values_changed' in diff:
        for path, change in diff['values_changed'].items():
            diff_data.append({
                'Path': format_path(path),
                'Change Type': 'Value Changed',
                'Old Value': format_value(change.get('old_value', '')),
                'New Value': format_value(change.get('new_value', ''))
            })
    if 'type_changes' in diff:
        for path, change in diff['type_changes'].items():
            diff_data.append({
                'Path': format_path(path),
                'Change Type': 'Type Changed',
                'Old Value': format_value(change.get('old_value', '')),
                'New Value': format_value(change.get('new_value', ''))
            })

    output_dir = 'schema'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    sanitized_account_code = "".join([c if c.isalnum() or c == '_' else "_" for c in data[i][0]])
    excel_file_path = f'{output_dir}/{sanitized_account_code}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Différences"

    ws.append(['Path', 'Change Type', 'Old Value', 'New Value'])

    
    for row in diff_data:
        ws.append([row['Path'], row['Change Type'], row['Old Value'], row['New Value']])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Sauvegarder le fichier Excel
    wb.save(excel_file_path)

    print(f"Differences exported to {excel_file_path}")