import csv
import json

from dotenv import load_dotenv
import openpyxl
from tabulate import tabulate 
from db.dbconn import SqlServerDB
from db.query import *
from openpyxl import Workbook
import xmltodict
from deepdiff import DeepDiff
import os
import datetime

load_dotenv()
db = SqlServerDB(os.getenv('SQL_SERVER_DB_SERVER'),os.getenv('SQL_SERVER_DB_DATABASE'),os.getenv('SQL_SERVER_DB_USERNAME'),os.getenv('SQL_SERVER_DB_PASSWORD'))
querys = Query()

def getAccountCode():
    allCode = []
    db.connect()
    data = db.execute_query(querys.accuntCode('bagsPAMF_CBS'))
    db.close
    for i in range(len(data)):
        allCode.append(data[i].get('Code'))
    return allCode

accountCode = getAccountCode()
db.connect()
wb = openpyxl.Workbook()
for i in range(len(accountCode)):
    
    old_dict = db.execute_query(querys.account('bagsPAMF_CBS',accountCode[i]))
    new_dict = db.execute_query(querys.account('bagsPAMF4',accountCode[i]))
    

    diff = DeepDiff(old_dict, new_dict)
    diff_data = []
    if 'values_changed' in diff:
        for path, change in diff['values_changed'].items():
            diff_data.append({
                'Path': path,
                'Change Type': 'Value Changed',
                'Old Value': change.get('old_value', ''),
                'New Value': change.get('new_value', '')
                })
    if 'type_changes' in diff:
        for path, change in diff['type_changes'].items():
            diff_data.append({
                'Path': path,
                'Change Type': 'Type Changed',
                'Old Value': change.get('old_value', ''),
                'New Value': change.get('new_value', '')
                })


    sanitized_account_code = "".join([c if c.isalnum() or c == '_' else "_" for c in accountCode[i]])
    ws = wb.create_sheet(title=sanitized_account_code)
    
    # 
    # ws = wb.active
    # ws.title = "Différences"

    ws.append(['Path', 'Change Type', 'Old Value', 'New Value'])

    
    for row in diff_data:
        ws.append([row['Path'], row['Change Type'], row['Old Value'], row['New Value']])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

if 'Sheet' in wb.sheetnames:
    default_sheet = wb['Sheet']
    if default_sheet.max_row == 1 and default_sheet.max_column == 1:
        wb.remove(default_sheet)


output_dir = 'out'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
excel_file_path = f'{output_dir}/account.xlsx'   
wb.save(excel_file_path)

print(f"Differences exported to {excel_file_path}")

db.close()